"""XLSX report export."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from takeoff_pro.data.models import Job
from takeoff_pro.estimate.pricing import price_job
from takeoff_pro.reports.data import build_takeoff_rows, job_summary_rows


def export_xlsx(job: Job, path: str | Path) -> Path:
    """Export an XLSX workbook report."""
    output_path = Path(path).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Job Summary"
    detail_sheet = workbook.create_sheet("Takeoff Detail")
    cost_sheet = workbook.create_sheet("Cost Detail")

    _write_rows(summary_sheet, [("Field", "Value"), *job_summary_rows(job)])
    _write_rows(
        detail_sheet,
        [
            ("Section", "Measurement", "Kind", "Quantity", "Unit", "Point Count"),
            *[
                (
                    row.section_name,
                    row.measurement_name,
                    row.kind,
                    row.quantity,
                    row.unit,
                    row.point_count,
                )
                for row in build_takeoff_rows(job)
            ],
        ],
    )
    _write_rows(
        cost_sheet,
        [
            ("Section", "Item ID", "Description", "Unit", "Quantity", "Unit Cost", "Total Cost"),
            *[
                (
                    line.section_name,
                    line.item_id,
                    line.description,
                    line.unit,
                    line.quantity,
                    line.unit_cost,
                    line.total_cost,
                )
                for line in price_job(job)
            ],
        ],
    )
    workbook.save(output_path)
    return output_path


def _write_rows(sheet: Worksheet, rows: list[tuple[object, ...]]) -> None:
    for row in rows:
        sheet.append(row)
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = column_cells[0].column_letter
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)

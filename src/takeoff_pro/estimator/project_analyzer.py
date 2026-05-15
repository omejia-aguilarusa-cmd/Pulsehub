"""
Project Analyzer — main orchestrator for the Senior Estimator analysis pipeline.

Reasoning sequence (applied to every project):
  1.  Identify all uploaded files.
  2.  Classify each file by document type and relevance.
  3.  Extract text from relevant files (PDF via PyMuPDF, XLSX via openpyxl).
  4.  Extract quantities (prefer takeoff > drawing > filename).
  5.  Extract pricing (ONLY from estimates — never from supplier quotes).
  6.  Apply methodology projections when source quantities are missing.
  7.  Normalize pricing per unit / BFT / GSF.
  8.  Compare against historical benchmark store.
  9.  Run QC checks and generate flags.
  10. Persist confirmed benchmarks to the store.
  11. Return a structured ProjectAnalysisReport.
"""

from __future__ import annotations

import logging
import re
from collections.abc import Callable
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pymupdf  # type: ignore[import-untyped]

from takeoff_pro.estimator.benchmark_store import BenchmarkStore, ProjectBenchmark, get_store
from takeoff_pro.estimator.document_classifier import (
    DocumentRelevance,
    DocumentType,
    classify_document,
)
from takeoff_pro.estimator.methodology import DRYWALL, PAINT, COVERAGE, SANITY, project_board_sf, project_paint_gallons
from takeoff_pro.estimator.pricing_extractor import ExtractedPricing, extract_pricing
from takeoff_pro.estimator.quantity_extractor import ExtractedQuantities, extract_quantities

LOGGER = logging.getLogger(__name__)

# ── Data structures ───────────────────────────────────────────────────────────

@dataclass
class DocumentRecord:
    path: Path
    doc_type: DocumentType
    relevance: DocumentRelevance
    classifier_confidence: float
    text_preview: str          # first 500 chars
    quantities: Optional[ExtractedQuantities]
    pricing: Optional[ExtractedPricing]
    reason: str                # Why classified this way


@dataclass
class QCFlag:
    severity: str       # "critical", "high", "medium", "low"
    category: str       # "Pricing", "Quantities", "Scope", "Sanity", "Files"
    description: str
    recommendation: str


@dataclass
class NormalizedMetrics:
    # Drywall
    drywall_per_unit: Optional[float] = None
    drywall_per_board_sf: Optional[float] = None
    drywall_per_gsf: Optional[float] = None
    drywall_labor_per_sheet: Optional[float] = None
    # Paint
    paint_per_unit: Optional[float] = None
    paint_per_gsf: Optional[float] = None
    # Combined
    total_per_unit: Optional[float] = None
    drywall_pct: Optional[float] = None
    paint_pct: Optional[float] = None
    # Ratios
    board_sf_per_unit: Optional[float] = None
    board_sf_per_gsf: Optional[float] = None
    paint_gal_per_unit: Optional[float] = None


@dataclass
class ProjectAnalysisReport:
    """Complete structured output from a project analysis run."""

    project_name: str

    # Document inventory
    documents: list[DocumentRecord]

    # Confirmed quantities (from takeoff, high confidence)
    units: Optional[int]
    gsf: Optional[float]
    nrsf: Optional[float]
    drywall_board_sf_with_waste: Optional[float]
    drywall_board_sf_before_waste: Optional[float]
    drywall_board_count_4x8: Optional[int]
    drywall_board_count_4x12: Optional[int]
    paint_gal_total: Optional[float]
    paint_wall_sf: Optional[float]
    paint_ceiling_sf: Optional[float]

    # Confirmed pricing (from estimate only)
    drywall_price: Optional[float]
    paint_price: Optional[float]
    total_price: Optional[float]
    drywall_labor_per_sheet: Optional[float]
    pricing_type: str

    # Scope flags
    paint_is_blended: bool       # interior + exterior not separated
    includes_exterior: bool
    includes_corridors: bool
    includes_amenities: bool

    # Methodology projections (when quantities not confirmed)
    projected_board_sf: Optional[float]
    projected_paint_gal: Optional[float]

    # Normalized metrics
    metrics: Optional[NormalizedMetrics]

    # Benchmark comparison
    benchmark_lines: list[str]

    # QC
    qc_flags: list[QCFlag]

    # Overall confidence
    confidence: str              # "high", "medium", "low"

    # Warnings from pricing extractor
    pricing_warnings: list[str]


# ── Analyzer ──────────────────────────────────────────────────────────────────

class ProjectAnalyzer:
    """
    Applies the Senior Construction Estimator methodology to every uploaded project.
    Retains and reuses confirmed benchmarks across sessions via BenchmarkStore.
    """

    def __init__(self, store: BenchmarkStore | None = None) -> None:
        self._store = store or get_store()

    def analyze(
        self,
        file_paths: list[Path],
        progress_callback: Callable[[str], None] | None = None,
    ) -> ProjectAnalysisReport:
        """Run the full analysis pipeline. Thread-safe — no Qt dependencies here."""
        n = len(file_paths)
        docs: list[DocumentRecord] = []

        # Step 1–3: classify and extract
        for i, path in enumerate(file_paths):
            if progress_callback:
                progress_callback(f"Reading {path.name} ({i + 1}/{n})…")
            docs.append(self._process_file(path))

        # Step 4–11: synthesize
        if progress_callback:
            progress_callback("Synthesizing quantities and pricing…")
        return self._synthesize(docs, file_paths)

    # ── File processing ───────────────────────────────────────────────────────

    def _process_file(self, path: Path) -> DocumentRecord:
        text = self._extract_text(path)
        doc_type, relevance, conf = classify_document(path.name, text)
        reason = f"Classified as {doc_type} (confidence={conf:.0%})"

        quantities: Optional[ExtractedQuantities] = None
        pricing: Optional[ExtractedPricing] = None

        if relevance in (DocumentRelevance.RELEVANT, DocumentRelevance.SUPPORTING):
            quantities = extract_quantities(text)
            # Only extract pricing from estimates and spreadsheets
            if doc_type in (
                DocumentType.ESTIMATE,
                DocumentType.PROPOSAL,
                DocumentType.PRICING_SPREADSHEET,
            ):
                pricing = extract_pricing(text)
            elif doc_type == DocumentType.SUPPLIER_QUOTE:
                reason += " — prices are supplier costs, not client prices"

        return DocumentRecord(
            path=path,
            doc_type=doc_type,
            relevance=relevance,
            classifier_confidence=conf,
            text_preview=text[:500].replace("\n", " "),
            quantities=quantities,
            pricing=pricing,
            reason=reason,
        )

    def _extract_text(self, path: Path) -> str:
        suffix = path.suffix.lower()
        try:
            if suffix == ".pdf":
                return self._pdf_text(path)
            if suffix in (".xlsx", ".xls"):
                return self._xlsx_text(path)
            if suffix in (".docx", ".doc"):
                return self._docx_text(path)
            if suffix in (".txt", ".csv"):
                return path.read_text(errors="replace")
        except Exception as exc:
            LOGGER.warning("Text extraction failed for %s: %s", path.name, exc)
        return ""

    def _pdf_text(self, path: Path) -> str:
        doc = pymupdf.open(str(path))
        try:
            return "\n".join(page.get_text("text") for page in doc)
        finally:
            doc.close()

    def _xlsx_text(self, path: Path) -> str:
        try:
            import openpyxl
        except ImportError:
            return ""
        wb = openpyxl.load_workbook(str(path), data_only=True)
        parts: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"[Sheet: {sheet_name}]")
            for row in ws.iter_rows():
                vals = [str(c.value) for c in row if c.value is not None]
                if vals:
                    parts.append(" | ".join(vals))
        return "\n".join(parts)

    def _docx_text(self, path: Path) -> str:
        try:
            import docx  # type: ignore[import-untyped]
            doc = docx.Document(str(path))
            return "\n".join(p.text for p in doc.paragraphs)
        except Exception:
            return ""

    # ── Synthesis ─────────────────────────────────────────────────────────────

    def _synthesize(
        self, docs: list[DocumentRecord], file_paths: list[Path]
    ) -> ProjectAnalysisReport:
        qc: list[QCFlag] = []
        pricing_warnings: list[str] = []

        # ── Project name ──────────────────────────────────────────────────────
        project_name = self._infer_project_name(docs, file_paths)

        # ── Select best documents by priority ────────────────────────────────
        estimate_docs = [d for d in docs if d.doc_type in (DocumentType.ESTIMATE, DocumentType.PROPOSAL) and d.pricing]
        takeoff_docs = [d for d in docs if d.doc_type in (DocumentType.TAKEOFF_REPORT, DocumentType.PRICING_SPREADSHEET) and d.quantities]
        drawing_docs = [d for d in docs if d.doc_type == DocumentType.DRAWING and d.quantities]

        # ── Quantities (takeoff > drawing > estimate) ─────────────────────────
        qty_sources = takeoff_docs + drawing_docs + estimate_docs
        units = self._best_int(qty_sources, "units")
        gsf = self._best_float(qty_sources, "gsf")
        nrsf = self._best_float(qty_sources, "nrsf")
        board_sf = self._best_float(takeoff_docs + estimate_docs, "drywall_board_sf_total")
        paint_gal = self._best_float(takeoff_docs + estimate_docs, "paint_gal_total")
        paint_wall_sf = self._best_float(takeoff_docs, "paint_wall_sf")
        paint_ceil_sf = self._best_float(takeoff_docs, "paint_ceiling_sf")

        # Derived board counts
        board_count_4x8 = int(board_sf / 32) if board_sf else None
        board_count_4x12 = int(board_sf / 48) if board_sf else None
        board_sf_before = board_sf / (1 + DRYWALL.waste_factor_board) if board_sf else None

        # ── Pricing (ONLY from estimate/proposal docs) ────────────────────────
        drywall_price: Optional[float] = None
        paint_price: Optional[float] = None
        total_price: Optional[float] = None
        drywall_labor_per_sheet: Optional[float] = None
        pricing_type = "unknown"
        paint_blended = False
        includes_exterior = False
        includes_corridors = False
        includes_amenities = False

        for doc in estimate_docs:
            p = doc.pricing
            if p is None:
                continue
            pricing_warnings.extend(p.warnings)
            if drywall_price is None and p.drywall_price:
                drywall_price = p.drywall_price
            if paint_price is None and p.paint_price:
                paint_price = p.paint_price
            if total_price is None and p.total_price:
                total_price = p.total_price
            if drywall_labor_per_sheet is None and p.drywall_labor_per_sheet:
                drywall_labor_per_sheet = p.drywall_labor_per_sheet
            if pricing_type == "unknown":
                pricing_type = p.pricing_type
            paint_blended = paint_blended or p.paint_is_blended
            includes_exterior = includes_exterior or p.paint_includes_exterior
            includes_corridors = includes_corridors or p.paint_includes_corridors
            includes_amenities = includes_amenities or p.paint_includes_amenities

        # Cross-check total
        if drywall_price and paint_price and not total_price:
            total_price = drywall_price + paint_price

        # ── Methodology projections ────────────────────────────────────────────
        projected_board_sf: Optional[float] = None
        projected_paint_gal: Optional[float] = None
        if units and gsf and not board_sf:
            common_gsf = gsf * DRYWALL.waste_factor_board   # simplified proxy
            res_gsf = gsf * 0.89
            nonres_gsf = gsf * 0.11
            proj = project_board_sf(units, res_gsf, nonres_gsf)
            projected_board_sf = proj["total"]

        if units and not paint_gal:
            unit_nrsf = (nrsf / units) if nrsf and units else 1026.0
            common_gsf = gsf * 0.11 if gsf else 0.0
            nonres_gsf = gsf * 0.04 if gsf else 0.0
            proj_gal = project_paint_gallons(units, unit_nrsf, common_gsf, nonres_gsf)
            projected_paint_gal = proj_gal["total_gal"]

        # ── Normalized metrics ────────────────────────────────────────────────
        metrics = self._normalize(
            units=units, gsf=gsf,
            board_sf=board_sf, paint_gal=paint_gal,
            drywall_price=drywall_price, paint_price=paint_price,
            total_price=total_price,
            drywall_labor_per_sheet=drywall_labor_per_sheet,
        )

        # ── QC checks ─────────────────────────────────────────────────────────
        qc.extend(self._run_qc_checks(
            docs=docs, units=units, gsf=gsf, board_sf=board_sf,
            drywall_price=drywall_price, paint_price=paint_price,
            metrics=metrics, paint_blended=paint_blended,
        ))

        # ── Benchmark comparison ───────────────────────────────────────────────
        benchmark_lines = self._compare_benchmarks(metrics)

        # ── Persist confirmed benchmarks ──────────────────────────────────────
        if drywall_price or paint_price:
            self._persist_benchmark(
                project_name=project_name,
                units=units or 0,
                gsf=gsf or 0.0,
                nrsf=nrsf or 0.0,
                board_sf=board_sf,
                paint_gal=paint_gal,
                drywall_price=drywall_price,
                paint_price=paint_price,
                total_price=total_price,
                drywall_labor_per_sheet=drywall_labor_per_sheet,
                pricing_type=pricing_type,
                includes_exterior=includes_exterior,
                docs=docs,
            )

        # ── Overall confidence ────────────────────────────────────────────────
        n_confirmed = sum(1 for v in [drywall_price, paint_price, board_sf, units] if v is not None)
        confidence = "high" if n_confirmed >= 3 else "medium" if n_confirmed >= 1 else "low"

        return ProjectAnalysisReport(
            project_name=project_name,
            documents=docs,
            units=units,
            gsf=gsf,
            nrsf=nrsf,
            drywall_board_sf_with_waste=board_sf,
            drywall_board_sf_before_waste=board_sf_before,
            drywall_board_count_4x8=board_count_4x8,
            drywall_board_count_4x12=board_count_4x12,
            paint_gal_total=paint_gal,
            paint_wall_sf=paint_wall_sf,
            paint_ceiling_sf=paint_ceil_sf,
            drywall_price=drywall_price,
            paint_price=paint_price,
            total_price=total_price,
            drywall_labor_per_sheet=drywall_labor_per_sheet,
            pricing_type=pricing_type,
            paint_is_blended=paint_blended,
            includes_exterior=includes_exterior,
            includes_corridors=includes_corridors,
            includes_amenities=includes_amenities,
            projected_board_sf=projected_board_sf,
            projected_paint_gal=projected_paint_gal,
            metrics=metrics,
            benchmark_lines=benchmark_lines,
            qc_flags=qc,
            confidence=confidence,
            pricing_warnings=pricing_warnings,
        )

    # ── Normalization ─────────────────────────────────────────────────────────

    def _normalize(
        self, *,
        units, gsf, board_sf, paint_gal,
        drywall_price, paint_price, total_price,
        drywall_labor_per_sheet,
    ) -> Optional[NormalizedMetrics]:
        if not units:
            return None
        m = NormalizedMetrics()
        if drywall_price:
            m.drywall_per_unit = round(drywall_price / units, 2)
            if board_sf:
                m.drywall_per_board_sf = round(drywall_price / board_sf, 4)
            if gsf:
                m.drywall_per_gsf = round(drywall_price / gsf, 4)
        m.drywall_labor_per_sheet = drywall_labor_per_sheet
        if paint_price:
            m.paint_per_unit = round(paint_price / units, 2)
            if gsf:
                m.paint_per_gsf = round(paint_price / gsf, 4)
        if total_price:
            m.total_per_unit = round(total_price / units, 2)
            if drywall_price:
                m.drywall_pct = round(drywall_price / total_price * 100, 1)
            if paint_price:
                m.paint_pct = round(paint_price / total_price * 100, 1)
        if board_sf:
            m.board_sf_per_unit = round(board_sf / units, 2)
            if gsf:
                m.board_sf_per_gsf = round(board_sf / gsf, 3)
        if paint_gal:
            m.paint_gal_per_unit = round(paint_gal / units, 2)
        return m

    # ── QC ────────────────────────────────────────────────────────────────────

    def _run_qc_checks(
        self, *, docs, units, gsf, board_sf, drywall_price,
        paint_price, metrics, paint_blended,
    ) -> list[QCFlag]:
        flags: list[QCFlag] = []

        if not drywall_price and not paint_price:
            flags.append(QCFlag(
                "critical", "Pricing",
                "No confirmed client-facing price found in any document.",
                "Locate signed estimate or proposal. Check if pricing spreadsheet has filled unit prices.",
            ))

        if not units:
            flags.append(QCFlag(
                "high", "Quantities",
                "Unit count could not be confirmed from any document.",
                "Check site plan title block, estimate header, or building program tab.",
            ))

        if not board_sf and units:
            flags.append(QCFlag(
                "medium", "Quantities",
                "Drywall board SF not confirmed. Methodology projection applied instead.",
                "Obtain digitized takeoff to tighten confidence from ±10% to ±2%.",
            ))

        if metrics and metrics.board_sf_per_unit is not None:
            bpu = metrics.board_sf_per_unit
            if bpu < SANITY.board_sf_per_unit_min:
                flags.append(QCFlag(
                    "high", "Sanity Check",
                    f"Board SF/unit = {bpu:,.0f} is BELOW the 4,500–5,500 target range. "
                    "May indicate missing double-layer assemblies or scope gap.",
                    "Verify that all party walls, corridor walls, and shaft walls are included.",
                ))
            elif bpu > SANITY.board_sf_per_unit_max:
                flags.append(QCFlag(
                    "medium", "Sanity Check",
                    f"Board SF/unit = {bpu:,.0f} is ABOVE the 4,500–5,500 target range. "
                    "May indicate double-counting or unusually complex framing.",
                    "Review wall-type legend and check for duplicate entries.",
                ))

        if metrics and metrics.paint_gal_per_unit is not None:
            gpu = metrics.paint_gal_per_unit
            if gpu < SANITY.paint_gal_per_unit_min:
                flags.append(QCFlag(
                    "medium", "Sanity Check",
                    f"Paint gal/unit = {gpu:.1f} is BELOW the 25–35 target. "
                    "May indicate missing primer coat, missing common areas, or low coat count.",
                    "Verify coat system: walls (3-coat), ceilings (2-coat), trim (3-coat).",
                ))
            elif gpu > SANITY.paint_gal_per_unit_max:
                flags.append(QCFlag(
                    "medium", "Sanity Check",
                    f"Paint gal/unit = {gpu:.1f} is ABOVE the 25–35 target. "
                    "If exterior paint is included, separate for accurate interior benchmark.",
                    "Confirm exterior scope and separate from interior paint gallonage.",
                ))

        if paint_blended:
            flags.append(QCFlag(
                "high", "Scope",
                "Paint price includes exterior scope — cannot confirm interior-only price. "
                "Do not use as interior paint benchmark without splitting.",
                "Request line-item breakdown: interior paint vs. exterior paint from GC or owner.",
            ))

        if metrics and metrics.drywall_per_unit is not None:
            dpu = metrics.drywall_per_unit
            if dpu < SANITY.drywall_per_unit_min:
                flags.append(QCFlag(
                    "medium", "Pricing",
                    f"Drywall $/unit = ${dpu:,.0f} is BELOW ${SANITY.drywall_per_unit_min:,.0f} "
                    "minimum. Verify scope — may be labor-only or incomplete.",
                    "Confirm pricing type: labor+material vs. labor-only.",
                ))

        if metrics and metrics.drywall_labor_per_sheet is not None:
            lps = metrics.drywall_labor_per_sheet
            if not (SANITY.drywall_labor_per_sheet_min <= lps <= SANITY.drywall_labor_per_sheet_max):
                flags.append(QCFlag(
                    "medium", "Pricing",
                    f"Labor/sheet = ${lps:.0f} is outside the ${SANITY.drywall_labor_per_sheet_min:.0f}–"
                    f"${SANITY.drywall_labor_per_sheet_max:.0f} expected range.",
                    "Confirm sheet size assumption (4×8 vs 4×12) and crew productivity.",
                ))

        # Check for ignored files that might be relevant
        ignored = [d for d in docs if d.relevance == DocumentRelevance.IGNORE
                   and d.doc_type not in (DocumentType.PAYROLL, DocumentType.FINANCIAL,
                                          DocumentType.LEGAL, DocumentType.UNRELATED)]
        for d in ignored:
            flags.append(QCFlag(
                "low", "Files",
                f"'{d.path.name}' was classified as {d.doc_type} and skipped.",
                "Open manually and verify it contains no pricing or quantity data.",
            ))

        return flags

    # ── Benchmark comparison ──────────────────────────────────────────────────

    def _compare_benchmarks(self, metrics: Optional[NormalizedMetrics]) -> list[str]:
        lines: list[str] = []
        if metrics is None:
            lines.append("No metrics available for benchmark comparison.")
            return lines

        if metrics.drywall_per_unit is not None:
            lines.append(
                f"Drywall $/unit: {self._store.compare('drywall_per_unit', metrics.drywall_per_unit)}"
            )

        if metrics.paint_per_unit is not None:
            lines.append(
                f"Paint $/unit: {self._store.compare('paint_per_unit', metrics.paint_per_unit)}"
            )

        if metrics.board_sf_per_unit is not None:
            lines.append(
                f"Board SF/unit: {self._store.compare('board_sf_per_unit', metrics.board_sf_per_unit)}"
            )

        avg_dw = self._store.avg_drywall_per_unit()
        avg_pt = self._store.avg_paint_per_unit()
        if avg_dw:
            lines.append(f"Historical avg drywall: ${avg_dw:,.0f}/unit "
                         f"(from {len(self._store.get_confirmed_drywall())} confirmed project(s))")
        if avg_pt:
            lines.append(f"Historical avg paint: ${avg_pt:,.0f}/unit "
                         f"(from {len(self._store.get_confirmed_paint())} confirmed project(s))")

        if not avg_dw and not avg_pt:
            lines.append("No confirmed pricing benchmarks yet in store.")

        return lines

    # ── Persistence ───────────────────────────────────────────────────────────

    def _persist_benchmark(
        self, *, project_name, units, gsf, nrsf, board_sf, paint_gal,
        drywall_price, paint_price, total_price, drywall_labor_per_sheet,
        pricing_type, includes_exterior, docs,
    ) -> None:
        dw_price_per_unit = drywall_price / units if drywall_price and units else None
        pt_price_per_unit = paint_price / units if paint_price and units else None
        dw_per_bsf = drywall_price / board_sf if drywall_price and board_sf else None
        n_confirmed = sum(1 for v in [drywall_price, paint_price, board_sf, units] if v is not None)
        confidence = "high" if n_confirmed >= 3 else "medium" if n_confirmed >= 1 else "low"

        bm = ProjectBenchmark(
            project_name=project_name,
            location="",
            units=units,
            gsf=gsf,
            nrsf=nrsf,
            drywall_board_sf_before_waste=board_sf / (1 + DRYWALL.waste_factor_board) if board_sf else None,
            drywall_board_sf_with_waste=board_sf,
            drywall_board_count_4x8=int(board_sf / 32) if board_sf else None,
            drywall_board_count_4x12=int(board_sf / 48) if board_sf else None,
            drywall_board_sf_per_unit=board_sf / units if board_sf and units else None,
            drywall_board_sf_per_gsf=board_sf / gsf if board_sf and gsf else None,
            drywall_price_turnkey=drywall_price,
            drywall_price_per_unit=dw_price_per_unit,
            drywall_price_per_board_sf=dw_per_bsf,
            drywall_labor_per_sheet=drywall_labor_per_sheet,
            paint_gal_total=paint_gal,
            paint_gal_per_unit=paint_gal / units if paint_gal and units else None,
            paint_price_turnkey=paint_price,
            paint_price_per_unit=pt_price_per_unit,
            paint_interior_exterior_split=not includes_exterior,
            total_price=total_price,
            total_price_per_unit=total_price / units if total_price and units else None,
            includes_exterior_paint=includes_exterior,
            pricing_type=pricing_type,
            source_files=[d.path.name for d in docs],
            confidence=confidence,
            notes="Auto-stored by ProjectAnalyzer.",
        )
        self._store.add(bm)

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _best_float(self, docs: list[DocumentRecord], field_name: str) -> Optional[float]:
        for d in docs:
            if d.quantities is None:
                continue
            val = getattr(d.quantities, field_name, None)
            if val is not None and val > 0:
                return float(val)
        return None

    def _best_int(self, docs: list[DocumentRecord], field_name: str) -> Optional[int]:
        v = self._best_float(docs, field_name)
        return int(v) if v is not None else None

    def _infer_project_name(self, docs: list[DocumentRecord], paths: list[Path]) -> str:
        # Try to find a project name in text
        patterns = [
            re.compile(r"project[:\s]+([A-Za-z0-9 &'\-]+(?:Apartments?|Apts?|Residences?|Condos?|Homes?|Towers?|Place|Grove|Park|Landing|Village|Commons|Heights))", re.IGNORECASE),
            re.compile(r"(?:estimate\s+for|proposal\s+for|bid\s+for)[:\s]+([A-Za-z0-9 &'\-]{5,60})", re.IGNORECASE),
        ]
        for d in docs:
            for p in patterns:
                m = p.search(d.text_preview)
                if m:
                    return m.group(1).strip()[:60]
        # Fall back to parent folder name
        if paths:
            return paths[0].parent.name or paths[0].stem
        return "Unknown Project"

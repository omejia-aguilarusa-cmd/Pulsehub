from __future__ import annotations

import csv
from pathlib import Path

import pymupdf
from openpyxl import load_workbook

from takeoff_pro.data import create_blank_job
from takeoff_pro.data.models import Job, Measurement, MeasurementKind, Point, TakeoffSection
from takeoff_pro.estimate.models import Assembly, AssemblyComponent, EstimateItem
from takeoff_pro.reports import export_csv, export_pdf, export_xlsx


def test_report_exports_create_csv_xlsx_and_pdf(tmp_path: Path) -> None:
    job = _priced_job()
    csv_path = export_csv(job, tmp_path / "report.csv")
    xlsx_path = export_xlsx(job, tmp_path / "report.xlsx")
    pdf_path = export_pdf(job, tmp_path / "report.pdf")

    with csv_path.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    assert rows[0]["row_type"] == "takeoff"
    assert rows[-1]["row_type"] == "cost"
    assert rows[-1]["total_cost"] == "130.0000"

    workbook = load_workbook(xlsx_path, read_only=True)
    assert workbook.sheetnames == ["Job Summary", "Takeoff Detail", "Cost Detail"]
    assert workbook["Job Summary"]["A1"].value == "Field"
    assert workbook["Cost Detail"]["G4"].value == 130
    workbook.close()

    document = pymupdf.open(pdf_path)
    assert document.page_count == 1
    page = document[0]
    text = page.get_text()
    pixmap = page.get_pixmap()
    document.close()

    assert "Priced Report" in text
    assert "Cost Detail" in text
    assert pixmap.width > 0
    assert pixmap.height > 0


def _priced_job() -> Job:
    job = create_blank_job("Priced")
    page_id = job.pages[0].id
    job.takeoff_sections = [
        TakeoffSection(
            id="direct",
            name="Baseboard",
            kind=MeasurementKind.LENGTH,
            estimate_reference_type="item",
            estimate_reference_id="WALL",
            measurements=[
                Measurement(
                    id="length",
                    name="Length",
                    kind=MeasurementKind.LENGTH,
                    page_id=page_id,
                    points=[Point(x=0, y=0), Point(x=10, y=0)],
                    quantity=10,
                    unit="LF",
                )
            ],
        ),
        TakeoffSection(
            id="assembly",
            name="Painted Wall",
            kind=MeasurementKind.AREA,
            estimate_reference_type="assembly",
            estimate_reference_id="PAINTED-WALL",
            measurements=[
                Measurement(
                    id="area",
                    name="Area",
                    kind=MeasurementKind.AREA,
                    page_id=page_id,
                    points=[
                        Point(x=0, y=0),
                        Point(x=10, y=0),
                        Point(x=10, y=10),
                        Point(x=0, y=10),
                    ],
                    quantity=100,
                    unit="SF",
                )
            ],
        ),
    ]
    job.items = [
        EstimateItem(item_id="WALL", description="Wall material", unit="LF", unit_cost=4.5),
        EstimateItem(item_id="PAINT", description="Paint", unit="GAL", unit_cost=35),
        EstimateItem(item_id="LABOR", description="Labor", unit="HR", unit_cost=65),
    ]
    job.assemblies = [
        Assembly(
            assembly_id="PAINTED-WALL",
            name="Painted wall",
            takeoff_unit="SF",
            components=[
                AssemblyComponent(item_id="PAINT", quantity_per_takeoff_unit=0.04),
                AssemblyComponent(item_id="LABOR", quantity_per_takeoff_unit=0.02),
            ],
        )
    ]
    return job
